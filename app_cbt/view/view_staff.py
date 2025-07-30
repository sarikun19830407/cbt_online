from django.shortcuts import render, redirect,get_object_or_404
from multiprocessing import context
from django.contrib.auth import authenticate, login
from django.contrib.auth.views import LoginView, LogoutView
from django.contrib.auth.decorators import login_required,user_passes_test,  permission_required
from django.views.decorators.csrf import csrf_protect
from django.conf import settings
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.contrib import messages
from django.contrib.auth.models import User
from django.http import HttpResponseForbidden
from django.db import transaction
from django.contrib.sessions.models import Session

from django.urls import reverse
import os
from django.http import FileResponse, HttpResponse, JsonResponse,Http404
import tempfile
from django.utils.timezone import now
from django.db import IntegrityError
import re
from django.views.decorators.cache import never_cache

from app_cbt import models
from app_cbt import forms_staff
from django.http import HttpResponse, Http404
import time 
from django.db.models import Count, Case, When, IntegerField
from app_cbt import forms
import xlrd
from django.core.files.storage import default_storage
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from io import BytesIO
from tempfile import NamedTemporaryFile
import openpyxl
from odf.opendocument import load
from odf.table import Table, TableRow, TableCell
from odf.text import P
from openpyxl import load_workbook
from django.db.models import Sum, Count, F
from openpyxl.styles import PatternFill, Font, Alignment
from django.utils import timezone
from django.template.loader import render_to_string
from weasyprint import HTML, CSS
import tempfile
import openpyxl
from openpyxl.utils import get_column_letter



def get_logged_in_users():
    sessions = Session.objects.filter(expire_date__gte=now())
    user_ids = []
    
    for session in sessions:
        data = session.get_decoded()
        if '_auth_user_id' in data:
            user_ids.append(data['_auth_user_id'])
    
    return models.Pengguna.objects.filter(id__in=user_ids, is_siswa=True)




@login_required(login_url=settings.LOGIN_URL)
@user_passes_test(lambda user: user.is_staff, login_url=settings.LOGIN_URL)
@csrf_protect
def staff(request):
    cari = request.POST.get('cari', request.GET.get('cari', '')) 
    items_per_page = int(request.GET.get('items_per_page', 10))

    users = get_logged_in_users()
    user_data = []

    for user in users:
        selesai = False
        kode_soal = models.SetingSoal.objects.filter(Nama_User=user).first()
        if kode_soal:
            total_soal = models.Soal_Siswa.objects.filter(Kode_Soal=kode_soal).count()
            total_jawab = models.Answer.objects.filter(Nama_User=user, Kode_Soal=kode_soal).exclude(Jawaban__isnull=True).count()
            selesai = total_soal > 0 and total_jawab >= total_soal

        user_data.append({
            'user': user,
            'selesai': selesai
        })

    if cari:
        user_data = [u for u in user_data if cari.lower() in u['user'].username.lower() or cari.lower() in u['user'].Nama.lower()]

    paginator = Paginator(user_data, items_per_page)
    page_number = request.GET.get('page', 1)  
    Data = paginator.get_page(page_number)
    jumlah = len(user_data)
    Data.start_index = (paginator.page(page_number).start_index() - 1)

    contex = {
        "data": "Home",
        "judul": "cbt-reset",
        "jumlah": jumlah,
        "JumlahPenggunaActive": models.Pengguna.objects.filter(is_active=True, is_siswa=True).count(),
        "JmlPengguna": models.Pengguna.objects.filter(is_siswa=True).count(),
        "Data": Data,
        "cari": cari,
        "items_per_page": items_per_page,
        "lembaga": "User telah Login",
        "placeholder": "Nama/User Name",
    }

    return render(request, 'super_admin/DasboardSuperuser.html', contex)


@csrf_protect
@login_required(login_url=settings.LOGIN_URL)
@user_passes_test(lambda u: u.is_staff, login_url=settings.LOGIN_URL)
def reset_login(request):
    if request.method == 'POST':
        user_ids = request.POST.getlist('user_ids')  # bisa lebih dari satu
        sessions = Session.objects.filter(expire_date__gte=now())

        reset_count = 0
        for session in sessions:
            data = session.get_decoded()
            if str(data.get('_auth_user_id')) in user_ids:
                session.delete()
                reset_count += 1

        messages.success(request, f"Sesi login {reset_count} pengguna berhasil direset.")
    return redirect('cbt:staff')



@login_required(login_url=settings.LOGIN_URL)
@user_passes_test(lambda user: user.is_staff, login_url=settings.LOGIN_URL)
@csrf_protect
def update_userlogin_Sataff(request):
    form = forms.FormUpdateSuperAdmin(request.POST or None, instance=request.user)
    if request.method == "POST":
        if form.is_valid():
            form.save()
            messages.add_message(request, messages.INFO, 'Data telah berhasil Tambahkan')
            return redirect(reverse('cbt:staff'))
        else:
            messages.error(request, 'Data Masih Salah.')
    context = {
        "data" : "Update user login",
        "NamaForm": "Form update user login",
        "judul":"CBT",
        "link":reverse("cbt:staff"),
        "form":form,
        'icon':"bi bi-pen"
        }
    return render (request, 'super_admin/form_user.html', context)






@login_required(login_url=settings.LOGIN_URL)
@user_passes_test(lambda user: user.is_staff, login_url=settings.LOGIN_URL)
@csrf_protect
def setting_soal(request):
    cari = request.POST.get('cari', request.GET.get('cari', ''))
    items_per_page = request.GET.get('items_per_page', '30')

    # Atur jumlah item per halaman
    if items_per_page == 'all':
        items_per_page = 1000000
    else:
        try:
            items_per_page = int(items_per_page)
        except (ValueError, TypeError):
            items_per_page = 30

    # ✅ Ambil semester aktif pertama (yang status=True)
    semester_obj = models.SEMESTER.objects.filter(status=True).first()

    # Ambil tahun pelajaran aktif
    tahun_aktif = models.TahunPelajaran.objects.filter(status=True).first()

    # Ambil data SetingSoal berdasarkan semester dan tahun pelajaran
    if semester_obj and tahun_aktif:
        data_list = models.SetingSoal.objects.filter(
            Semester=semester_obj.Semester,
            Tahun_Pelajaran=tahun_aktif
        )
    else:
        data_list = models.SetingSoal.objects.none()

    # Filter pencarian jika keyword 'cari' diisi
    if cari:
        data_list = data_list.filter(Mapel__Nama_Mapel__icontains=cari)

    # Proses pagination
    paginator = Paginator(data_list, items_per_page)
    page_number = request.GET.get('page', 1)

    try:
        Data = paginator.page(page_number)
    except PageNotAnInteger:
        Data = paginator.page(1)
    except EmptyPage:
        Data = paginator.page(paginator.num_pages)

    # Hitung index dan jumlah total data
    Data.start_index = (Data.start_index() - 1)
    jumlah = data_list.count()

    context = {
        "data": f"Soal Semester {semester_obj.Semester if semester_obj else '-'} TP.{tahun_aktif if tahun_aktif else '-'}",
        "judul": "CBT-Setting",
        "Data": Data,
        "icon": "bi bi-database-check",
        "semester": semester_obj,
        "jumlah": jumlah,
        "cari": cari,
        "items_per_page": str(items_per_page) if items_per_page != 1000000 else 'all',
        "lembaga": "Mapel Kelas",
        "placeholder": "Mapel",
    }

    return render(request, 'staff/seting_soal.html', context)





@login_required(login_url=settings.LOGIN_URL)
@user_passes_test(lambda user: user.is_staff, login_url=settings.LOGIN_URL)
def aktifkan_soal(request, kode_soal):
    soal = get_object_or_404(models.SetingSoal, Kode_Soal=kode_soal)

    if soal.aktif:
        # Jika soal ini sedang aktif → maka nonaktifkan
        soal.aktif = False
        soal.save()
        messages.warning(request, f"Soal Kelas {soal.Kelas} | Mapel {soal.Mapel} telah dinonaktifkan.")
    else:
        # Cek apakah sudah ada soal lain yang aktif
        soal_aktif_lain = models.SetingSoal.objects.filter(aktif=True).exclude(Kode_Soal=soal.Kode_Soal).first()

        if soal_aktif_lain:
            messages.error(
                request,
                f"Tidak bisa mengaktifkan soal ini karena Mapel {soal_aktif_lain.Mapel} Kelas {soal_aktif_lain.Kelas} sedang aktif."
            )
        else:
            # Aktifkan soal karena tidak ada soal lain yang aktif
            soal.aktif = True
            soal.waktu_aktif = timezone.now()
            soal.save()
            messages.success(request, f"Soal Kelas {soal.Kelas} | Mapel {soal.Mapel} berhasil diaktifkan.")

    return redirect('cbt:setting_soal')






@login_required(login_url=settings.LOGIN_URL)
@user_passes_test(lambda user: user.is_staff, login_url=settings.LOGIN_URL)
@csrf_protect
def buat_seting_soal(request):
    form = forms_staff.SetingSoalForm(request.POST or None)
    if request.method == "POST":
        if form.is_valid():
            form.instance.Nama_User = request.user
            form.instance.Nama_Lembaga = request.user.Nama_Lembaga
            form.instance.arsip_soal = True
            semester_instance = get_object_or_404(models.SEMESTER, status=True)
            form.instance.Semester = semester_instance
            tahun_aktif = get_object_or_404(models.TahunPelajaran, status=True)
            form.instance.Tahun_Pelajaran = tahun_aktif
            form.save()
            messages.add_message(request, messages.INFO, 'Data telah berhasil Tambahkan')
            return redirect(reverse('cbt:setting_soal'))
        else:
            messages.error(request, 'Data Masih Salah.')
    context = {
        "data" : "Tambah Setting Soal",
        "NamaForm": "Form Setting Soal",
        "judul":"CBT-Setting",
        "link":reverse("cbt:setting_soal"),
        "form":form,
        'icon':'bi bi-file-plus'
        
        }
    return render (request, 'staff/form.html', context)




@login_required(login_url=settings.LOGIN_URL)
@user_passes_test(lambda user: user.is_staff, login_url=settings.LOGIN_URL)
@csrf_protect
def hapus_setting_soal(request):
    if request.method == "POST":
        selected_ids = request.POST.getlist('selected_ids')
        if not selected_ids:
            messages.warning(request, 'Tidak ada data yang dipilih untuk dihapus.')
            return redirect(reverse("cbt:setting_soal"))  
        
        Data = models.SetingSoal.objects.filter(id__in=selected_ids)
        if not Data.exists():
            raise Http404("Data Madrasah tidak ditemukan.")
        
        # Pemeriksaan kepemilikan data
        for data in Data:
            if data.Nama_User != request.user:
                messages.error(request, "Anda tidak memiliki hak akses untuk menghapus data ini.")
                return redirect(reverse("cbt:setting_soal"))
        
        if 'confirm' in request.POST:
            Data.delete()  
            messages.success(request, 'Data telah berhasil dihapus.')
            return redirect(reverse("cbt:setting_soal"))
        context = {
            "data" : f"Hapus {Data.count()} mapel",
            "NamaForm": "Form Hapus Seting Soal",
            "judul":"PPDB Hapus Seting Soal",
            "link":reverse("cbt:setting_soal"),
            "hapus": reverse('cbt:hapus_setting_soal'),
            "Hapus": f"{Data.count()} Mapel",
            "ket":"Madrasah",
            "Data": Data,
            "icon":"bi bi-trash3"
            }
        return render (request, 'staff/hapus_data.html', context)
    return redirect(reverse("cbt:setting_soal"))



@login_required(login_url=settings.LOGIN_URL)
@user_passes_test(lambda user: user.is_staff, login_url=settings.LOGIN_URL)
@csrf_protect
def Ubah_setting_soal (request, pk):
    try:
        Data = get_object_or_404(models.SetingSoal, id=pk)
    except Http404:
        messages.error(request, "Data pengguna tidak ditemukan.")
        return redirect(reverse('cbt:setting_soal'))  # Redirect ke halaman daftar pengguna
    
    if request.user.is_staff:
        # Pengguna provinsi hanya bisa mengedit dirinya sendiri
        if Data.Nama_User != request.user:
            messages.error(request, "Anda hanya bisa mengedit data Anda sendiri.")
            return redirect(reverse('cbt:setting_soal'))
    form = forms_staff.SetingSoalForm(request.POST or None, instance=Data)
    if request.method == "POST":
        if form.is_valid():
            form.save()
            messages.add_message(request, messages.INFO, 'Data telah berhasil diubah')
            return redirect(reverse('cbt:setting_soal'))
        else:
            messages.error(request, 'Data Masih Salah.')
    context = {
        "data" : f"Ubah Soal {Data.Mapel}",
        "NamaForm": "Form ubah Setting Soal",
        "judul":"CBT-Setting Soal",
        "link":reverse("cbt:setting_soal"),
        "form":form,
        'icon':"bi bi-pen"
        }
    return render (request, 'staff/form.html', context)









@login_required(login_url=settings.LOGIN_URL)
@user_passes_test(lambda user: user.is_staff, login_url=settings.LOGIN_URL)
@csrf_protect
def buat_soal_siswa(request, pk):
    try:
        setting = get_object_or_404(models.SetingSoal, id=pk)
    except Http404:
        messages.error(request, "Halaman tidak ditemukan.")
        return redirect(reverse('cbt:setting_soal'))  # Redirect ke halaman daftar pengguna
    
    if request.user.is_staff:
        if setting.Nama_User != request.user:
            messages.error(request, "Anda hanya bisa mengedit data Anda sendiri.")
            return redirect(reverse('cbt:setting_soal'))

    # Pastikan hanya user yang sesuai yang bisa akses
    if request.user.is_staff and setting.Nama_User != request.user:
        messages.error(request, "Anda hanya bisa mengedit data Anda sendiri.")
        return redirect(reverse('cbt:setting_soal'))

    # Hitung nomor soal berikutnya berdasarkan soal terakhir
    last_soal = models.Soal_Siswa.objects.filter(Kode_Soal=setting).order_by('-Nomor').first()
    if last_soal and last_soal.Nomor and str(last_soal.Nomor).isdigit():
        next_nomor = str(int(last_soal.Nomor) + 1)
    else:
        next_nomor = '1'

    # Handle form POST (pengiriman soal baru)
    if request.method == 'POST':
        form = forms_staff.SoalSiswaForm(request.POST)
        if form.is_valid():
            soal = form.save(commit=False)
            soal.Nama_User = request.user
            soal.Kode_Soal = setting
            soal.Kelas = setting.Kelas
            soal.Mapel = setting.Mapel
            soal.Nomor = next_nomor  # Set nomor otomatis
            soal.save()

            messages.success(request, 'Soal berhasil ditambahkan.')
            return redirect(reverse('cbt:buat_soal_siswa', kwargs={'pk': pk}))
        else:
            messages.error(request, 'Terjadi kesalahan dalam form.')
    else:
        # Inisialisasi form saat pertama kali buka
        initial_data = {
            'Kode_Soal': setting,
            'Mapel': setting.Mapel,
            'Kelas': setting.Kelas,
            
        }
        form = forms_staff.SoalSiswaForm(initial=initial_data)

    # Ambil daftar soal yang sudah ada
    soal_list = models.Soal_Siswa.objects.filter(Kode_Soal=setting).order_by('Nomor')

    # Kirim data ke template
    context = {
        "data": f"Buat Soal Mapel {setting.Mapel} Nomor : {next_nomor}",
        "NamaForm": "Form Buat Soal Siswa",
        "judul": "CBT - Buat Soal",
        "link": reverse("cbt:setting_soal"),
        "form": form,
        'setting': setting,
        'soal_list': soal_list,
        "next_nomor": next_nomor,
        "icon": "bi bi-file-earmark-plus",
    }

    return render(request, 'staff/buat_soal_siswa.html', context)


@login_required(login_url=settings.LOGIN_URL)
@user_passes_test(lambda user: user.is_staff, login_url=settings.LOGIN_URL)
@csrf_protect
def hapus_soal (request, pk):
    next_url = request.GET.get('next') or reverse('cbt:setting_soal')
    try:
        Data = get_object_or_404(models.Soal_Siswa, id=pk)
    except Http404:
        messages.error(request, "Data soal tidak ditemukan.")
        return redirect(next_url)
    
    if request.user.is_staff:
        # Pengguna provinsi hanya bisa mengedit dirinya sendiri
        if Data.Nama_User != request.user:
            messages.error(request, "Anda hanya bisa mengedit data Anda sendiri.")
            return redirect(next_url)
        
    next_url = reverse('cbt:buat_soal_siswa', kwargs={'pk': Data.Kode_Soal.pk})
    if request.method == "POST":
        Data.delete()
        messages.add_message(request, messages.INFO, 'Data telah berhasil Hapus')
        return redirect(reverse('cbt:buat_soal_siswa', kwargs={'pk': Data.Kode_Soal.pk}))
    context = {
        "data" : f"Hapus Soal No : {Data}",
        "judul":"CBT-user staff",
        "link": next_url,
        "Data":f"Hapus Soal No : {Data}",
        "icon": "bi bi-trash3"
        }
    return render (request, 'staff/hapus.html', context)




@login_required(login_url=settings.LOGIN_URL)
@user_passes_test(lambda user: user.is_staff, login_url=settings.LOGIN_URL)
@csrf_protect
def Ubah_soal(request, pk):
    next_url = request.GET.get('next') or reverse('cbt:setting_soal')
    try:
        Data = get_object_or_404(models.Soal_Siswa, id=pk)
    except Http404:
        messages.error(request, "Data soal tidak ditemukan.")
        return redirect(next_url)

    # Pastikan user hanya bisa edit soal miliknya
    if request.user.is_staff and Data.Nama_User != request.user:
        messages.error(request, "Anda hanya bisa mengedit soal Anda sendiri.")
        return redirect(next_url)

    # Ambil next URL (default ke setting soal jika tidak ada)
    
    
    form = forms_staff.SoalSiswaForm(request.POST or None, instance=Data)
    if request.method == "POST":
        if form.is_valid():
            form.save()
            messages.success(request, 'Soal berhasil diperbarui.')
            return redirect(next_url)
        else:
            messages.error(request, 'Form tidak valid.')

    context = {
        "data": f"Ubah soal nomor: {Data.Nomor}",
        "NamaForm": "Form Ubah Soal",
        "judul": "CBT - Ubah Soal",
        "link": next_url,
        "form": form,
        "soal_list": models.Soal_Siswa.objects.filter(Kode_Soal=Data.Kode_Soal).order_by('Nomor'),
        "setting" : Data.Kode_Soal
    }
    return render(request, 'staff/buat_soal_siswa.html', context)




def clear_upload_session_data(request):
    keys = [
        'duplicate_users',
        'invalid_usernames',
    ]
    for key in keys:
        request.session.pop(key, None)






def get_text_content(element):
    result = []
    for node in element.childNodes:
        if hasattr(node, 'data'):
            result.append(str(node.data))
        else:
            result.append(get_text_content(node))  # rekursif jika nested
    return ''.join(result)




@login_required(login_url=settings.LOGIN_URL)
@user_passes_test(lambda user: user.is_staff, login_url=settings.LOGIN_URL)
@csrf_protect
@login_required
def upload_soal_excel(request, pk):
    setting = get_object_or_404(models.SetingSoal, pk=pk)
    form = forms.UploadFormSoal(request.POST or None, request.FILES or None)

    success_count = 0
    invalid_rows = []

    if request.method == 'POST' and form.is_valid():
        excel_file = form.cleaned_data['file_excel']
        file_name = excel_file.name.lower()

        # Simpan file sementara
        temp_file_path = default_storage.save('temp/' + file_name, excel_file)

        try:
            data_rows = []

            # Deteksi tipe file
            if file_name.endswith('.xlsx') or file_name.endswith('.xlsm'):
                wb = load_workbook(filename=os.path.join(settings.MEDIA_ROOT, temp_file_path))
                ws = wb.active
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i == 0:
                        continue
                    data_rows.append([str(cell).strip() if cell else '' for cell in row])

            elif file_name.endswith('.xls'):
                wb = xlrd.open_workbook(os.path.join(settings.MEDIA_ROOT, temp_file_path))
                sheet = wb.sheet_by_index(0)
                for i in range(1, sheet.nrows):
                    row = [str(sheet.cell_value(i, j)).strip() for j in range(sheet.ncols)]
                    data_rows.append(row)

            elif file_name.endswith('.ods'):
                doc = load(os.path.join(settings.MEDIA_ROOT, temp_file_path))
                sheets = doc.spreadsheet.getElementsByType(Table)
                if not sheets:
                    raise Exception("Tidak ada sheet ditemukan di file .ods")
                sheet = sheets[0]

                skip_first = True
                for row in sheet.getElementsByType(TableRow):
                    if skip_first:
                        skip_first = False
                        continue

                    values = []
                    for cell in row.getElementsByType(TableCell):
                        texts = cell.getElementsByType(P)
                        text_content = ''.join([
                            get_text_content(p) for p in texts
                        ])
                        repeat = int(cell.getAttribute("numbercolumnsrepeated") or 1)
                        values.extend([text_content] * repeat)

                    data_rows.append(values)

            else:
                messages.error(request, "Format file tidak didukung.")
                return redirect(request.path)

            # Bersihkan baris kosong
            data_rows = [row for row in data_rows if any(cell.strip() for cell in row)]

            # Proses baris
            for idx, row in enumerate(data_rows, start=2):
                if len(row) < 7:
                    invalid_rows.append({'row': idx, 'error': 'Baris tidak lengkap, minimal 7 kolom.'})
                    continue

                nomor = row[0].strip()
                soal = row[1].strip()
                opsi_a = row[2].strip()
                opsi_b = row[3].strip()
                opsi_c = row[4].strip()
                opsi_d = row[5].strip()
                kunci = row[6].strip().upper()
                nilai = row[7]

                if kunci not in ['A', 'B', 'C', 'D']:
                    invalid_rows.append({'row': idx, 'error': f'Kunci tidak valid: {kunci}'})
                    continue

                try:
                    models.Soal_Siswa.objects.create(
                        Nama_User=request.user,
                        Kode_Soal=setting,
                        Mapel=setting.Mapel,
                        Kelas=setting.Kelas,
                        Nomor=nomor,
                        Soal=soal,
                        A=opsi_a,
                        B=opsi_b,
                        C=opsi_c,
                        D=opsi_d,
                        Kunci_Jawaban=kunci,
                        Nilai= nilai,
                    )
                    success_count += 1
                except Exception as e:
                    invalid_rows.append({'row': idx, 'error': f'Gagal insert soal: {str(e)}'})

            # Pesan hasil
            if invalid_rows:
                request.session['invalid_rows'] = invalid_rows
                messages.warning(request, f"{success_count} soal berhasil diunggah. {len(invalid_rows)} gagal.")
            else:
                messages.success(request, f"{success_count} soal berhasil diunggah!")

            return redirect(reverse('cbt:buat_soal_siswa', kwargs={'pk': pk}))

        except Exception as e:
            messages.error(request, f"Kesalahan saat membaca file: {str(e)}")

    context = {
        "data": f"Upload Soal {setting.Mapel}|Kelas {setting.Kelas}",
        "NamaForm": "Form Upload Soal",
        "judul": "Upload Soal Siswa",
        "form": form,
        "success_count": success_count,
        "link": 'cbt:setingsoal',  # sesuaikan dengan nama URL back
        "icon": "bi bi-upload",
        "link":reverse('cbt:buat_soal_siswa', kwargs={'pk': pk})
    }
    return render(request, "staff/upload.html", context)







@login_required(login_url=settings.LOGIN_URL)
@user_passes_test(lambda user: user.is_staff, login_url=settings.LOGIN_URL)
@csrf_protect
def download_template_soal(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Template Soal"

    # Styling
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
    header_font = Font(bold=True)
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Header
    headers = ['Nomor', 'Soal (HTML)', 'A', 'B', 'C', 'D', 'Kunci', 'Nilai']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = align_center
        ws.column_dimensions[cell.column_letter].width = 30

    # Contoh baris isi
    ws.append([
        '1',
        'contoh soal',
        'opsi 1', 'opsi 2', 'opsi 3', 'opsi 4',
        'A',"Nilai 1-10"
        
    ])

    # Border untuk isi
    for row in ws.iter_rows(min_row=2, max_row=2, min_col=1, max_col=8):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Simpan ke BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=template_soal.xlsx'
    return response






@login_required(login_url=settings.LOGIN_URL)
@user_passes_test(lambda user: user.is_staff, login_url=settings.LOGIN_URL)
@csrf_protect
def nilai_setingsoal(request, kode_soal):
    seting_soal = get_object_or_404(models.SetingSoal, id=kode_soal)

    kelas = seting_soal.Kelas
    mapel = seting_soal.Mapel
    lembaga = seting_soal.Nama_Lembaga

    # Ambil semua Rombel yang relevan
    rombels = models.Answer.objects.filter(
        Kode_Soal=seting_soal
    ).values_list('Rombel', flat=True).distinct()

    # Hitung total nilai maksimal dari semua soal
    total_nilai_maksimal = sum(
        soal.Nilai or 0
        for soal in models.Soal_Siswa.objects.filter(Kode_Soal=seting_soal)
    )

    jawaban = models.Answer.objects.filter(
        Kode_Soal=seting_soal
    ).select_related("Nama_User", "Nomor_Soal")

    nilai_siswa = {}
    for ans in jawaban:
        user = ans.Nama_User
        if user.id not in nilai_siswa:
            nama = user.get_full_name() or getattr(user, 'Nama', None) or user.username
            nilai_siswa[user.id] = {
                "nama": nama,
                "rombel": getattr(ans, 'Rombel', None),
                "total_nilai": 0,
                "total_benar": 0,
                "total_soal": 0
            }

        if ans.Jawaban is not None:
            nilai_siswa[user.id]["total_soal"] += 1

        if ans.Jawaban_Benar:
            nilai_soal = ans.Nomor_Soal.Nilai or 0
            nilai_siswa[user.id]["total_nilai"] += nilai_soal
            nilai_siswa[user.id]["total_benar"] += 1

    for data in nilai_siswa.values():
        total_soal = data['total_soal']
        data["persentase"] = round((data["total_benar"] / total_soal) * 100, 2) if total_soal > 0 else 0
        data["nilai_akhir"] = round((data["total_nilai"] / total_nilai_maksimal) * 100, 2) if total_nilai_maksimal > 0 else 0

    context = {
        "data": f"Daftar Nilai : {mapel} | Kelas : {kelas}",
        "NamaForm": "Hasil Ujian",
        "seting_soal": seting_soal,
        "total_nilai_maksimal": total_nilai_maksimal,
        "nilai_siswa": sorted(nilai_siswa.values(), key=lambda x: x["nama"]),
    }

    return render(request, "staff/nilai_setingsoal.html", context)





@login_required(login_url=settings.LOGIN_URL)
@user_passes_test(lambda user: user.is_staff, login_url=settings.LOGIN_URL)
def export_nilai_excel_setting(request, kode_soal):
    seting_soal = get_object_or_404(models.SetingSoal, Kode_Soal=kode_soal)

    jawaban = models.Answer.objects.filter(
        Kode_Soal=seting_soal
    ).select_related("Nama_User", "Nomor_Soal", "Rombel")

    soal_list = list(models.Soal_Siswa.objects.filter(Kode_Soal=seting_soal))
    total_nilai_maksimal = sum(soal.Nilai or 0 for soal in soal_list)

    nilai_siswa = {}
    for ans in jawaban:
        user = ans.Nama_User
        if user.id not in nilai_siswa:
            nama = user.get_full_name() or getattr(user, 'Nama', None) or user.username
            rombel_nama = str(getattr(ans, 'Rombel', ''))
            nilai_siswa[user.id] = {
                "nama": nama,
                "rombel": rombel_nama,
                "total_nilai": 0,
                "total_benar": 0,
                "total_soal": 0
            }

        if ans.Jawaban is not None:
            nilai_siswa[user.id]["total_soal"] += 1

        if ans.Jawaban_Benar:
            nilai_soal = ans.Nomor_Soal.Nilai or 0
            nilai_siswa[user.id]["total_nilai"] += nilai_soal
            nilai_siswa[user.id]["total_benar"] += 1

    for data in nilai_siswa.values():
        total_soal = data['total_soal']
        data["persentase"] = round((data["total_benar"] / total_soal) * 100, 2) if total_soal > 0 else 0
        data["nilai_akhir"] = round((data["total_nilai"] / total_nilai_maksimal) * 100, 2) if total_nilai_maksimal > 0 else 0

    # Buat workbook dan sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Nilai - {seting_soal.Kode_Soal}"

    # Header kolom
    ws.append(["No", "Nama", "Rombel", "Total Benar", "Total Soal",  "Nilai Akhir"])

    # Data siswa
    for idx, data in enumerate(sorted(nilai_siswa.values(), key=lambda x: x["nama"]), start=1):
        ws.append([
            idx,
            data["nama"],
            data["rombel"],
            data["total_benar"],
            data["total_soal"],
            data["nilai_akhir"]
        ])

    # Set lebar kolom otomatis
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    nama_file = f"Kelas_{seting_soal.Kelas}_{seting_soal.Mapel}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{nama_file}"'
    wb.save(response)
    return response







@login_required(login_url=settings.LOGIN_URL)
@user_passes_test(lambda user: user.is_staff, login_url=settings.LOGIN_URL)
@csrf_protect
def lihat_soal_siswa(request, pk):
    Data = get_object_or_404(models.SetingSoal, pk=pk)
    
    cari = request.POST.get('cari', request.GET.get('cari', ''))
    items_per_page = request.GET.get('items_per_page', '30')
    
    if items_per_page == 'all':
        items_per_page = 1000000
    else:
        try:
            items_per_page = int(items_per_page)
        except (ValueError, TypeError):
            items_per_page = 30

    # Filter berdasarkan Data soal
    data_list = models.Soal_Siswa.objects.filter(Nomor=Data).order_by('Nomor')

    if cari:
        data_list = data_list.filter(Soal__icontains=cari)  # cari di isi soal

    paginator = Paginator(data_list, items_per_page)
    page_number = request.GET.get('page', 1)

    try:
        Data = paginator.page(page_number)
    except PageNotAnInteger:
        Data = paginator.page(1)
    except EmptyPage:
        Data = paginator.page(paginator.num_pages)

    semua = str(items_per_page) if items_per_page != 1000000 else 'all'
    jumlah = data_list.count()
    
    context = {
        "data": "Lihat Soal",
        "judul": f"Soal untuk: ",
        "Data": Data,
        "jumlah": jumlah,
        "cari": cari,
        "items_per_page": semua,
        "lembaga": "Soal Siswa",
        "placeholder": "Nomor",
        "icon": "bi bi-list-ul",
        "Data": Data,
        "link":reverse("cbt:setting_soal"),
    }

    return render(request, 'staff/list_soal_siswa.html', context)






@login_required(login_url=settings.LOGIN_URL)
@user_passes_test(lambda user: user.is_staff, login_url=settings.LOGIN_URL)
@csrf_protect
def hapus_soal_pilih(request):
    if request.method == "POST":
        selected_ids = request.POST.getlist('selected_ids')
        if not selected_ids:
            messages.warning(request, 'Tidak ada data yang dipilih untuk dihapus.')
            return redirect(reverse('cbt:lihat_soal_siswa', kwargs={'pk': kode_soal_pk})) 
        
        Data = models.Soal_Siswa.objects.filter(id__in=selected_ids)
        if not Data.exists():
            raise Http404("Data Soal tidak ditemukan.")

        for data in Data:
            if data.Nama_User != request.user:
                messages.error(request, "Anda tidak memiliki hak akses untuk menghapus data ini.")
                return redirect(reverse('cbt:lihat_soal_siswa', kwargs={'pk': kode_soal_pk}))

        # Dapatkan pk soal dari salah satu data
        kode_soal_pk = Data.first().Kode_Soal.pk

        if 'confirm' in request.POST:
            Data.delete()
            messages.success(request, 'Data telah berhasil dihapus.')
            return redirect(reverse('cbt:lihat_soal_siswa', kwargs={'pk': kode_soal_pk}))

        context = {
            "data": f"Hapus {Data.count()} soal",
            "NamaForm": "Form Hapus Soal",
            "judul": "CBT - Hapus Soal",
            "link": reverse("cbt:lihat_soal_siswa", kwargs={"pk": kode_soal_pk}),  # tombol kembali
            "hapus": reverse('cbt:hapus_soal_pilih'),
            "Data": Data,
            "Hapus": f"{Data.count()} soal",
            "ket": "Soal",
            "icon": "bi bi-trash3"
        }
        return render(request, 'staff/hapus_data.html', context)

    return redirect(reverse('cbt:lihat_soal_siswa', kwargs={'pk': kode_soal_pk}))









# views.py

@login_required(login_url=settings.LOGIN_URL)
@user_passes_test(lambda user: user.is_staff, login_url=settings.LOGIN_URL)
@csrf_protect
def ujicoba_soal(request, pk):
    setting = get_object_or_404(models.SetingSoal, pk=pk)
    soal_list = models.Soal_Siswa.objects.filter(Kode_Soal=setting).order_by('Nomor')

    if request.method == 'POST':
        jawaban = {}
        for soal in soal_list:
            user_jawaban = request.POST.get(f"jawaban_{soal.id}", "")
            jawaban[str(soal.id)] = user_jawaban
        request.session['ujicoba_jawaban'] = jawaban
        return redirect('cbt:ujicoba_selesai', pk=pk)

    return render(request, 'staff/ujicoba_soal.html', {
        'setting': setting,
        'soal_list': soal_list
    })



@login_required(login_url=settings.LOGIN_URL)
@user_passes_test(lambda user: user.is_staff, login_url=settings.LOGIN_URL)
@csrf_protect
def ujicoba_selesai(request, pk):
    setting = get_object_or_404(models.SetingSoal, pk=pk)
    soal_list = models.Soal_Siswa.objects.filter(Kode_Soal=setting).order_by('Nomor')
    jawaban = request.session.get('ujicoba_jawaban', {})

    # Hitung skor (opsional)
    benar = 0
    for soal in soal_list:
        if jawaban.get(str(soal.id)) == soal.Kunci_Jawaban:
            benar += 1
    total = soal_list.count()
    skor = int((benar / total) * 100) if total > 0 else 0

    # Hapus session setelah selesai
    del request.session['ujicoba_jawaban']

    return render(request, 'staff/ujicoba_selesai.html', {
        'setting': setting,
        'skor': skor,
        'total': total,
        'benar': benar,
    })



@login_required(login_url=settings.LOGIN_URL)
@user_passes_test(lambda user: user.is_staff, login_url=settings.LOGIN_URL)
@csrf_protect
def lihat_soal_siswa(request, pk):
    Data = get_object_or_404(models.SetingSoal, pk=pk)
    
    cari = request.POST.get('cari', request.GET.get('cari', ''))
    items_per_page = request.GET.get('items_per_page', '30')
    
    if items_per_page == 'all':
        items_per_page = 1000000
    else:
        try:
            items_per_page = int(items_per_page)
        except (ValueError, TypeError):
            items_per_page = 30

    

    # Filter berdasarkan Data soal
    data_list = models.Soal_Siswa.objects.filter(Kode_Soal=Data).order_by('Nomor')

    if cari:
        data_list = data_list.filter(Nomor__icontains=cari)  # cari di isi soal

    paginator = Paginator(data_list, items_per_page)
    page_number = request.GET.get('page', 1)

    try:
        Data = paginator.page(page_number)
    except PageNotAnInteger:
        Data = paginator.page(1)
    except EmptyPage:
        Data = paginator.page(paginator.num_pages)

    semua = str(items_per_page) if items_per_page != 1000000 else 'all'
    jumlah = data_list.count()
    
    context = {
        "data": "Lihat Soal",
        "judul": f"Soal untuk: ",
        "Data": Data,
        "jumlah": jumlah,
        "cari": cari,
        "items_per_page": semua,
        "lembaga": "Soal Siswa",
        "placeholder": "Nomor",
        "icon": "bi bi-list-ul",
        "Data": Data,
        "link":reverse("cbt:setting_soal"),
    }

    return render(request, 'staff/list_soal_siswa.html', context)


@login_required(login_url=settings.LOGIN_URL)
@user_passes_test(lambda user: user.is_staff, login_url=settings.LOGIN_URL)
@csrf_protect
def jawaban_siswa(request, id_daftarnilai):
    try:
        daftar_nilai = get_object_or_404(models.DaftarNilai, id=id_daftarnilai)
    except Http404:
        messages.error(request, "Analisis tidak ditemukan.")
        return redirect(reverse('cbt:daftar_nilai'))  
    
    kelas = daftar_nilai.Kelas
    rombel = daftar_nilai.Rombel
    mapel = daftar_nilai.Mapel
    semester = daftar_nilai.Semester
    tahun_pelajaran = daftar_nilai.Tahun_Pelajaran
    lembaga = daftar_nilai.Nama_Lembaga

    # Ambil seting soal sesuai kriteria
    seting_soal = models.SetingSoal.objects.filter(
        Nama_Lembaga=lembaga,
        Kelas=kelas,
        Mapel=mapel,
        Semester=semester,
        Tahun_Pelajaran=tahun_pelajaran
    ).first()

    if not seting_soal:
        messages.warning(request, "Seting Soal tidak ditemukan.")
        return redirect(reverse('cbt:daftar_nilai'))

    # Ambil semua soal untuk seting soal ini
    soal_list = models.Soal_Siswa.objects.filter(
        Kode_Soal=seting_soal
    ).order_by('Nomor')
    total_soal = soal_list.count()
    total_nilai_maksimal = sum(soal.Nilai for soal in soal_list if soal.Nilai)

    # Ambil semua siswa (user.id) dari DaftarNilai sesuai kelas & rombel
    siswa_list = models.Answer.objects.filter(
        Kode_Soal=seting_soal,
        Kelas=kelas,
        Rombel=rombel
    ).select_related("Nama_User", "Nomor_Soal")

    jawaban_siswa = {}
    for siswa in siswa_list:
        user = siswa.Nama_User
        nama_siswa = user.get_full_name() or getattr(user, 'Nama', None) or user.username
        
        jawaban_siswa[user.id] = {
            'nama': nama_siswa,
            'jawaban': {},
            'total_nilai': 0,
            'total_benar': 0
        }

    # Ambil semua jawaban dari siswa-siswa tsb
    answers = models.Answer.objects.filter(
        Kode_Soal=seting_soal,
        Kelas=kelas,
        Rombel=rombel
    ).select_related('Nama_User', 'Nomor_Soal')

    for answer in answers:
        user_id = answer.Nama_User.id
        if user_id in jawaban_siswa:
            jawaban_siswa[user_id]['jawaban'][answer.Nomor_Soal.Nomor] = {
                'huruf': answer.Jawaban,
                'benar': answer.Jawaban_Benar,
                'nilai': answer.Nilai_Siswa or 0
            }

            if answer.Jawaban_Benar:
                jawaban_siswa[user_id]['total_nilai'] += answer.Nilai_Siswa or 0
                jawaban_siswa[user_id]['total_benar'] += 1

    # Hitung persentase
    for user_data in jawaban_siswa.values():
        user_data['persentase'] = (user_data['total_benar'] / total_soal) * 100 if total_soal > 0 else 0

    # Statistik soal
    statistik_soal = []
    for soal in soal_list:
        total_jawaban = models.Answer.objects.filter(
            Nomor_Soal=soal,
            Kelas=kelas,
            Rombel=rombel
        ).count()
        
        benar = models.Answer.objects.filter(
            Nomor_Soal=soal,
            Jawaban_Benar=True,
            Kelas=kelas,
            Rombel=rombel
        ).count()
        
        persentase = (benar / total_jawaban) * 100 if total_jawaban > 0 else 0

        statistik_soal.append({
            'nomor': soal.Nomor,
            'kunci': soal.Kunci_Jawaban,
            'benar': benar,
            'total': total_jawaban,
            'persentase': round(persentase, 1),
            'nilai_soal': soal.Nilai or 0
        })

    context = {
        "data": f"Analisis : {mapel} Kelas : {kelas}/{rombel}",
        "judul": "CBT-Analisis soal",
        'daftar_nilai': daftar_nilai,
        'seting_soal': seting_soal,
        'soal_list': soal_list,
        'jawaban_siswa': sorted(jawaban_siswa.values(), key=lambda x: x['nama']),
        'statistik_soal': statistik_soal,
        'total_nilai_maksimal': total_nilai_maksimal
    }
    return render(request, 'staff/jawaban_siswa.html', context)





@login_required(login_url=settings.LOGIN_URL)
@user_passes_test(lambda user: user.is_staff, login_url=settings.LOGIN_URL)
@csrf_protect
def export_jawaban_siswa_excel(request, id_daftarnilai):
    try:
        daftar_nilai = get_object_or_404(models.DaftarNilai, id=id_daftarnilai)
    except Http404:
        messages.error(request, "Analisis tidak ditemukan.")
        return redirect(reverse('cbt:daftar_nilai'))  
    
    kelas = daftar_nilai.Kelas
    rombel = daftar_nilai.Rombel
    mapel = daftar_nilai.Mapel
    semester = daftar_nilai.Semester
    tahun_pelajaran = daftar_nilai.Tahun_Pelajaran
    lembaga = daftar_nilai.Nama_Lembaga

    # Ambil seting soal sesuai kriteria
    seting_soal = models.SetingSoal.objects.filter(
        Nama_Lembaga=lembaga,
        Kelas=kelas,
        Mapel=mapel,
        Semester=semester,
        Tahun_Pelajaran=tahun_pelajaran
    ).first()

    if not seting_soal:
        messages.warning(request, "Seting Soal tidak ditemukan.")
        return redirect(reverse('cbt:daftar_nilai'))

    # Ambil semua soal untuk seting soal ini
    soal_list = models.Soal_Siswa.objects.filter(
        Kode_Soal=seting_soal
    ).order_by('Nomor')
    total_soal = soal_list.count()
    total_nilai_maksimal = sum(soal.Nilai for soal in soal_list if soal.Nilai)

    # Ambil semua siswa (user.id) dari DaftarNilai sesuai kelas & rombel
    siswa_list = models.Answer.objects.filter(
        Kode_Soal=seting_soal,
        Kelas=kelas,
        Rombel=rombel
    ).select_related("Nama_User", "Nomor_Soal")

    jawaban_siswa = {}
    for siswa in siswa_list:
        user = siswa.Nama_User
        nama_siswa = user.get_full_name() or getattr(user, 'Nama', None) or user.username
        
        jawaban_siswa[user.id] = {
            'nama': nama_siswa,
            'jawaban': {},
            'total_nilai': 0,
            'total_benar': 0
        }

    # Ambil semua jawaban dari siswa-siswa tsb
    answers = models.Answer.objects.filter(
        Kode_Soal=seting_soal,
        Kelas=kelas,
        Rombel=rombel
    ).select_related('Nama_User', 'Nomor_Soal')

    for answer in answers:
        user_id = answer.Nama_User.id
        if user_id in jawaban_siswa:
            jawaban_siswa[user_id]['jawaban'][answer.Nomor_Soal.Nomor] = {
                'huruf': answer.Jawaban,
                'benar': answer.Jawaban_Benar,
                'nilai': answer.Nilai_Siswa or 0
            }

            if answer.Jawaban_Benar:
                jawaban_siswa[user_id]['total_nilai'] += answer.Nilai_Siswa or 0
                jawaban_siswa[user_id]['total_benar'] += 1

    # Hitung persentase
    for user_data in jawaban_siswa.values():
        user_data['persentase'] = (user_data['total_benar'] / total_soal) * 100 if total_soal > 0 else 0


    # Buat workbook Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Analisis Siswa"

    # Header
    headers = ['No', 'Nama Siswa'] + [f"Soal {soal.Nomor}" for soal in soal_list] + ['Jawaban Benar', 'Persentase']
    ws.append(headers)

    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = header_fill

    # Data siswa
    benar_per_soal = {soal.Nomor: 0 for soal in soal_list}
    for i, siswa in enumerate(sorted(jawaban_siswa.values(), key=lambda x: x['nama']), start=1):
        row = [i, siswa['nama']]
        for soal in soal_list:
            nomor = soal.Nomor
            data = siswa['jawaban'].get(nomor, {})
            huruf = data.get('huruf', '-')
            benar = data.get('benar', False)
            if benar:
                benar_per_soal[nomor] += 1
            row.append(huruf)
        row.append(siswa['total_benar'])
        persentase = (siswa['total_benar'] / soal_list.count()) * 100 if soal_list else 0
        row.append(f"{persentase:.1f}%")

        # Tambahkan baris ke sheet
        ws.append(row)

        # Warna sel jawaban benar
        for j, soal in enumerate(soal_list, start=3):  # Kolom jawaban mulai dari kolom C
            jawaban = siswa['jawaban'].get(soal.Nomor)
            if jawaban and jawaban['benar']:
                cell = ws.cell(row=ws.max_row, column=j)
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    # Tambah baris total
    total_row = ['Total Benar', '']
    for soal in soal_list:
        total_row.append(benar_per_soal[soal.Nomor])
    total_row.append('')
    total_row.append(f"{len(jawaban_siswa)} siswa")

    ws.append(total_row)

    # Style untuk baris total
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    # Auto width
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[col_letter].width = adjusted_width

    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"Analisis_{mapel}_{kelas}_{rombel}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'

    wb.save(response)
    return response


@login_required(login_url=settings.LOGIN_URL)
@user_passes_test(lambda user: user.is_staff, login_url=settings.LOGIN_URL)
@csrf_protect
def export_jawaban_siswa_pdf(request, id_daftarnilai):
    try:
        daftar_nilai = get_object_or_404(models.DaftarNilai, id=id_daftarnilai)
    except Http404:
        messages.error(request, "Analisis tidak ditemukan.")
        return redirect(reverse('cbt:daftar_nilai'))  
    
    kelas = daftar_nilai.Kelas
    rombel = daftar_nilai.Rombel
    mapel = daftar_nilai.Mapel
    semester = daftar_nilai.Semester
    tahun_pelajaran = daftar_nilai.Tahun_Pelajaran
    lembaga = daftar_nilai.Nama_Lembaga

    # Ambil seting soal sesuai kriteria
    seting_soal = models.SetingSoal.objects.filter(
        Nama_Lembaga=lembaga,
        Kelas=kelas,
        Mapel=mapel,
        Semester=semester,
        Tahun_Pelajaran=tahun_pelajaran
    ).first()

    if not seting_soal:
        messages.warning(request, "Seting Soal tidak ditemukan.")
        return redirect(reverse('cbt:daftar_nilai'))

    # Ambil semua soal untuk seting soal ini
    soal_list = models.Soal_Siswa.objects.filter(
        Kode_Soal=seting_soal
    ).order_by('Nomor')
    total_soal = soal_list.count()
    total_nilai_maksimal = sum(soal.Nilai for soal in soal_list if soal.Nilai)

    # Ambil semua siswa (user.id) dari DaftarNilai sesuai kelas & rombel
    siswa_list = models.Answer.objects.filter(
        Kode_Soal=seting_soal,
        Kelas=kelas,
        Rombel=rombel
    ).select_related("Nama_User", "Nomor_Soal")

    jawaban_siswa = {}
    for siswa in siswa_list:
        user = siswa.Nama_User
        nama_siswa = user.get_full_name() or getattr(user, 'Nama', None) or user.username
        
        jawaban_siswa[user.id] = {
            'nama': nama_siswa,
            'jawaban': {},
            'total_nilai': 0,
            'total_benar': 0
        }

    # Ambil semua jawaban dari siswa-siswa tsb
    answers = models.Answer.objects.filter(
        Kode_Soal=seting_soal,
        Kelas=kelas,
        Rombel=rombel
    ).select_related('Nama_User', 'Nomor_Soal')

    for answer in answers:
        user_id = answer.Nama_User.id
        if user_id in jawaban_siswa:
            jawaban_siswa[user_id]['jawaban'][answer.Nomor_Soal.Nomor] = {
                'huruf': answer.Jawaban,
                'benar': answer.Jawaban_Benar,
                'nilai': answer.Nilai_Siswa or 0
            }

            if answer.Jawaban_Benar:
                jawaban_siswa[user_id]['total_nilai'] += answer.Nilai_Siswa or 0
                jawaban_siswa[user_id]['total_benar'] += 1

    # Hitung persentase
    for user_data in jawaban_siswa.values():
        user_data['persentase'] = (user_data['total_benar'] / total_soal) * 100 if total_soal > 0 else 0

    # Statistik soal
    statistik_soal = []
    for soal in soal_list:
        total_jawaban = models.Answer.objects.filter(
            Nomor_Soal=soal,
            Kelas=kelas,
            Rombel=rombel
        ).count()
        
        benar = models.Answer.objects.filter(
            Nomor_Soal=soal,
            Jawaban_Benar=True,
            Kelas=kelas,
            Rombel=rombel
        ).count()
        
        persentase = (benar / total_jawaban) * 100 if total_jawaban > 0 else 0
        
        statistik_soal.append({
            'nomor': soal.Nomor,
            'kunci': soal.Kunci_Jawaban,
            'benar': benar,
            'total': total_jawaban,
            'persentase': round(persentase, 1),
            'nilai_soal': soal.Nilai or 0
        })

    html_string = render_to_string("staff/Jawaban_siswa_pdf.html", {
        "data": f"Analisis : {mapel} Kelas : {kelas}/{rombel}",
        "judul": "CBT-Analisis soal",
        'daftar_nilai': daftar_nilai,
        'seting_soal': seting_soal,
        'soal_list': soal_list,
        'jawaban_siswa': sorted(jawaban_siswa.values(), key=lambda x: x['nama']),
        'statistik_soal': statistik_soal,
        'total_nilai_maksimal': total_nilai_maksimal,
        "lembaga":lembaga,
        "tahun_pelajaran":tahun_pelajaran,
        
    })


    response = HttpResponse(content_type='application/pdf')
    filename = f"Analisis_{mapel}_{kelas}_{rombel}.pdf"
    response['Content-Disposition'] = f'filename="{filename}"'

    # Konversi cm ke points (1 cm = 28.35 points)
    width = 8.5 * 28.35  # Lebar 8.5 cm
    height = 13 * 28.35   # Tinggi 13 cm

    pdf_options = {
        'page-size': 'legal',  # Default, akan di-override oleh page-width dan page-height
        'page-width': f'{width}pt',
        'page-height': f'{height}pt',
        'orientation': 'Landscape',
        'margin-top': '0.5cm',
        'margin-right': '0.5cm',
        'margin-bottom': '0.5cm',
        'margin-left': '0.5cm',
        'encoding': "UTF-8",
    }

    try:
        HTML(string=html_string).write_pdf(
            target=response,
            options=pdf_options
        )
    except Exception as e:
        messages.error(request, f"Gagal membuat PDF: {str(e)}")
        return redirect('cbt:daftar_nilai_view', id_daftarnilai=id_daftarnilai)

    return response



@login_required(login_url=settings.LOGIN_URL)
@user_passes_test(lambda user: user.is_staff, login_url=settings.LOGIN_URL)
@csrf_protect
def daftar_nilai(request):
    cari = request.POST.get('cari', request.GET.get('cari', ''))
    items_per_page = request.GET.get('items_per_page', '30')
    
    if items_per_page == 'all':
        items_per_page = 1000000
    else:
        try:
            items_per_page = int(items_per_page)
        except (ValueError, TypeError):
            items_per_page = 30

    # ✅ Ambil semester aktif pertama (status=True)
    semester_obj = models.SEMESTER.objects.filter(status=True).first()
    semester = semester_obj.Semester if semester_obj else None

    # ✅ Ambil tahun pelajaran aktif
    tahun_aktif = models.TahunPelajaran.objects.filter(status=True).first()

    if semester_obj and tahun_aktif:
        data_list = models.DaftarNilai.objects.filter(
            Semester=semester_obj.Semester,
            Tahun_Pelajaran=tahun_aktif
        )
    else:
        data_list = models.DaftarNilai.objects.none()


    # ✅ Filter pencarian berdasarkan kelas
    if cari:
        data_list = data_list.filter(Kelas__Kelas__icontains=cari)

    # ✅ Paginasi
    paginator = Paginator(data_list, items_per_page)
    page_number = request.GET.get('page', 1)

    try:
        Data = paginator.page(page_number)
    except PageNotAnInteger:
        Data = paginator.page(1)
    except EmptyPage:
        Data = paginator.page(paginator.num_pages)

    semua = str(items_per_page) if items_per_page != 1000000 else 'all'
    jumlah = data_list.count()
    
    context = {
        "data": f"Daftar Nilai Semester {semester if semester else '-'} TP {tahun_aktif if tahun_aktif else '-'}",
        "judul": "Daftar Nilai",
        "Data": Data,
        "jumlah": jumlah,
        "cari": cari,
        "semester":  semester,
        "items_per_page": semua,
        "lembaga": "Daftar Nilai",
        "placeholder": "Kelas",
        "icon": "bi bi-list-ul",
        "link": reverse("cbt:setting_soal"),
    }

    return render(request, 'staff/list_daftar_nilai.html', context)



@login_required(login_url=settings.LOGIN_URL)
@user_passes_test(lambda user: user.is_staff, login_url=settings.LOGIN_URL)
@csrf_protect
def tambah_daftar_nilai(request):
    form = forms_staff.Daftar_nilai(request.POST or None)
    if request.method == "POST":
        if form.is_valid():
            form.instance.Nama_User = request.user
            form.instance.Nama_Lembaga = request.user.Nama_Lembaga
            semester_instance = get_object_or_404(models.SEMESTER, status=True)
            form.instance.Semester = semester_instance
            tahun_aktif = get_object_or_404(models.TahunPelajaran, status=True)
            form.instance.Tahun_Pelajaran = tahun_aktif
            form.save()
            messages.add_message(request, messages.INFO, 'Data telah berhasil Tambahkan')
            return redirect(reverse('cbt:daftar_nilai'))
        else:
            messages.error(request, 'Data Masih Salah.')
    context = {
        "data" : "Tambah daftar nilai",
        "NamaForm": "Form Tambah daftar nilai",
        "judul":"CBT-daftar nilai",
        "link":reverse("cbt:daftar_nilai"),
        "form":form,
        "icon":"bi bi-plus-circle"
        }
    return render (request, 'super_admin/form.html', context)







@login_required(login_url=settings.LOGIN_URL)
@user_passes_test(lambda user: user.is_staff, login_url=settings.LOGIN_URL)
@csrf_protect
def ubah_daftar_nilai (request, pk):
    try:
        Data = get_object_or_404(models.DaftarNilai, id=pk)
    except Http404:
        messages.error(request, "Data pengguna tidak ditemukan.")
        return redirect(reverse('cbt:daftar_nilai'))  # Redirect ke halaman daftar pengguna
    
    if request.user.is_staff:
        # Pengguna provinsi hanya bisa mengedit dirinya sendiri
        if Data.Nama_User != request.user:
            messages.error(request, "Anda hanya bisa mengedit data Anda sendiri.")
            return redirect(reverse('cbt:daftar_nilai'))
    form = forms_staff.Daftar_nilai(request.POST or None, instance=Data)
    if request.method == "POST":
        if form.is_valid():
            form.save()
            messages.add_message(request, messages.INFO, 'Data telah berhasil diubah')
            return redirect(reverse('cbt:daftar_nilai'))
        else:
            messages.error(request, 'Data Masih Salah.')
    context = {
        "data" : f"Ubah Soal {Data.Mapel}",
        "NamaForm": "Form ubah Setting Soal",
        "judul":"CBT-Setting Soal",
        "link":reverse("cbt:daftar_nilai"),
        "form":form,
        'icon':"bi bi-pen"
        }
    return render (request, 'staff/form.html', context)



@login_required(login_url=settings.LOGIN_URL)
@user_passes_test(lambda user: user.is_staff, login_url=settings.LOGIN_URL)
@csrf_protect
def hapus_daftar_nilai(request):
    
    if request.method == "POST":
        selected_ids = request.POST.getlist('selected_ids')
        if not selected_ids:
            messages.warning(request, 'Tidak ada data yang dipilih untuk dihapus.')
            return redirect(reverse('cbt:daftar_nilai'))  
        
        Data = models.DaftarNilai.objects.filter(id__in=selected_ids)
        if not Data.exists():
            raise Http404("Data Madrasah tidak ditemukan.")
        
        # Pemeriksaan kepemilikan data
        for data in Data:
            if data.Nama_User != request.user:
                messages.error(request, "Anda tidak memiliki hak akses untuk menghapus data ini.")
                return redirect(reverse('cbt:daftar_nilai'))
        
        if 'confirm' in request.POST:
            Data.delete()  
            messages.success(request, 'Data telah berhasil dihapus.')
            return redirect(reverse('cbt:daftar_nilai')) 
        
        context = {
            "data" : f"Hapus {Data.count()} siswa",
            "NamaForm": "Form Hapus Madrsah",
            "judul":"PPDB Hapus Madrsah",
            "link":reverse("cbt:daftar_nilai"),
            "hapus": reverse('cbt:hapus_daftar_nilai'),
            "Data": Data,
            "Hapus":f"{Data.count()} siswa",
            "ket":"Madrasah",
            "icon":"bi bi-trash3"
        }
        return render(request, 'super_admin/hapu_data.html', context)
    
    return redirect(reverse('cbt:daftar_nilai'))











@login_required(login_url=settings.LOGIN_URL)
@user_passes_test(lambda user: user.is_staff, login_url=settings.LOGIN_URL)
@csrf_protect
def daftar_nilai_view(request, id_daftarnilai):
    try:
        daftar_nilai = get_object_or_404(models.DaftarNilai, id=id_daftarnilai)
    except Http404:
        messages.error(request, "Daftar nilai tidak ditemukan.")
        return redirect(reverse('cbt:daftar_nilai'))

    # Hanya user yang membuat data yang bisa mengakses
    if request.user.is_staff and daftar_nilai.Nama_User != request.user:
        messages.error(request, "Anda hanya bisa mengedit data Anda sendiri.")
        return redirect(reverse('cbt:daftar_nilai'))

    # Ambil data dasar
    kelas = daftar_nilai.Kelas
    rombel = daftar_nilai.Rombel
    mapel = daftar_nilai.Mapel
    lembaga = daftar_nilai.Nama_Lembaga
    semester = daftar_nilai.Semester
    tahun_pelajaran = daftar_nilai.Tahun_Pelajaran

    # Ambil data SetingSoal berdasarkan semua parameter penting
    seting_soal = models.SetingSoal.objects.filter(
        Nama_Lembaga=lembaga,
        Kelas=kelas,
        Mapel=mapel,
        Semester=semester,
        Tahun_Pelajaran=tahun_pelajaran
    ).first()

    if not seting_soal:
        messages.error(request, "Seting Soal tidak ditemukan.")
        return redirect("cbt:daftar_nilai")

    # Hitung total nilai maksimal
    soal_list = models.Soal_Siswa.objects.filter(Kode_Soal=seting_soal)
    total_nilai_maksimal = sum(soal.Nilai or 0 for soal in soal_list)

    # Ambil jawaban siswa berdasarkan soal dan filter
    jawaban = models.Answer.objects.filter(
        Kode_Soal=seting_soal,
        Kelas=kelas,
        Rombel=rombel
    ).select_related("Nama_User", "Nomor_Soal")

    nilai_siswa = {}
    for ans in jawaban:
        user = ans.Nama_User
        if user.id not in nilai_siswa:
            nama = user.get_full_name() or getattr(user, 'Nama', None) or user.username
            nilai_siswa[user.id] = {
                "nama": nama,
                "total_nilai": 0,
                "total_benar": 0,
                "total_soal": 0
            }

        # Jika ada jawaban, tambah total soal
        if ans.Jawaban is not None:
            nilai_siswa[user.id]["total_soal"] += 1

        # Jika benar, tambah nilai sesuai bobot soal
        if ans.Jawaban_Benar:
            nilai_soal = ans.Nomor_Soal.Nilai or 0
            nilai_siswa[user.id]["total_nilai"] += nilai_soal
            nilai_siswa[user.id]["total_benar"] += 1

    # Hitung persentase dan nilai akhir
    for data in nilai_siswa.values():
        total_soal = data['total_soal']
        data["persentase"] = round((data["total_benar"] / total_soal) * 100, 2) if total_soal > 0 else 0
        data["nilai_akhir"] = round((data["total_nilai"] / total_nilai_maksimal) * 100, 2) if total_nilai_maksimal > 0 else 0

    # Kirim ke template
    context = {
        "judul": "CBT - Daftar Nilai",
        "data": f"Daftar Nilai: {mapel} {kelas} / {rombel}",
        "NamaForm": "Form Tambah Daftar Nilai",
        "seting_soal": seting_soal,
        "total_nilai_maksimal": total_nilai_maksimal,
        "nilai_siswa": sorted(nilai_siswa.values(), key=lambda x: x["nama"]),
    }

    return render(request, "staff/daftar_nilai.html", context)







@login_required(login_url=settings.LOGIN_URL)
@user_passes_test(lambda user: user.is_staff, login_url=settings.LOGIN_URL)
@csrf_protect
def export_daftar_nilai_excel(request, id_daftarnilai):
    try:
        daftar_nilai = get_object_or_404(models.DaftarNilai, id=id_daftarnilai)
    except Http404:
        messages.error(request, "Daftar nilai tidak ditemukan.")
        return redirect(reverse('cbt:daftar_nilai'))

    if request.user.is_staff and daftar_nilai.Nama_User != request.user:
        messages.error(request, "Anda hanya bisa mengunduh data Anda sendiri.")
        return redirect('cbt:daftar_nilai')

    # Ambil data dasar
    kelas = daftar_nilai.Kelas
    rombel = daftar_nilai.Rombel
    mapel = daftar_nilai.Mapel
    lembaga = daftar_nilai.Nama_Lembaga
    semester = daftar_nilai.Semester
    tahun_pelajaran = daftar_nilai.Tahun_Pelajaran

    # Ambil data SetingSoal berdasarkan semua parameter penting
    seting_soal = models.SetingSoal.objects.filter(
        Nama_Lembaga=lembaga,
        Kelas=kelas,
        Mapel=mapel,
        Semester=semester,
        Tahun_Pelajaran=tahun_pelajaran
    ).first()

    if not seting_soal:
        messages.error(request, "Seting Soal tidak ditemukan.")
        return redirect("cbt:daftar_nilai")

    # Hitung total nilai maksimal
    soal_list = models.Soal_Siswa.objects.filter(Kode_Soal=seting_soal)
    total_nilai_maksimal = sum(soal.Nilai or 0 for soal in soal_list)

    # Ambil jawaban siswa berdasarkan soal dan filter
    jawaban = models.Answer.objects.filter(
        Kode_Soal=seting_soal,
        Kelas=kelas,
        Rombel=rombel
    ).select_related("Nama_User", "Nomor_Soal")

    nilai_siswa = {}
    for ans in jawaban:
        user = ans.Nama_User
        if user.id not in nilai_siswa:
            nama = user.get_full_name() or getattr(user, 'Nama', None) or user.username
            nilai_siswa[user.id] = {
                "nama": nama,
                "total_nilai": 0,
                "total_benar": 0,
                "total_soal": 0
            }

        # Jika ada jawaban, tambah total soal
        if ans.Jawaban is not None:
            nilai_siswa[user.id]["total_soal"] += 1

        # Jika benar, tambah nilai sesuai bobot soal
        if ans.Jawaban_Benar:
            nilai_soal = ans.Nomor_Soal.Nilai or 0
            nilai_siswa[user.id]["total_nilai"] += nilai_soal
            nilai_siswa[user.id]["total_benar"] += 1

    # Hitung persentase dan nilai akhir
    for data in nilai_siswa.values():
        total_soal = data['total_soal']
        data["persentase"] = round((data["total_benar"] / total_soal) * 100, 2) if total_soal > 0 else 0
        data["nilai_akhir"] = round((data["total_nilai"] / total_nilai_maksimal) * 100, 2) if total_nilai_maksimal > 0 else 0

    # Buat Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Daftar Nilai"

    # Header
    headers = ["No", "Nama Siswa", "Benar", "Skor Akhir (100)"]
    ws.append(headers)

    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = header_fill

    # Data baris siswa
    for i, data in enumerate(sorted(nilai_siswa.values(), key=lambda x: x['nama']), start=1):
        ws.append([
            i,
            data['nama'],
            data['total_benar'],
            data['nilai_akhir']
        ])

    # Auto width
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    # Output Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"Nilai_{mapel}_{kelas}_{rombel}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'

    wb.save(response)
    return response




@login_required(login_url=settings.LOGIN_URL)
@user_passes_test(lambda user: user.is_staff, login_url=settings.LOGIN_URL)
@csrf_protect
def export_daftar_nilai_pdf(request, id_daftarnilai):
    daftar_nilai = get_object_or_404(models.DaftarNilai, id=id_daftarnilai)

    if request.user.is_staff and daftar_nilai.Nama_User != request.user:
        messages.error(request, "Anda hanya bisa mengunduh data Anda sendiri.")
        return redirect('cbt:daftar_nilai')

    # Ambil data dasar
    kelas = daftar_nilai.Kelas
    rombel = daftar_nilai.Rombel
    mapel = daftar_nilai.Mapel
    lembaga = daftar_nilai.Nama_Lembaga
    semester = daftar_nilai.Semester
    tahun_pelajaran = daftar_nilai.Tahun_Pelajaran

    # Ambil data SetingSoal berdasarkan semua parameter penting
    seting_soal = models.SetingSoal.objects.filter(
        Nama_Lembaga=lembaga,
        Kelas=kelas,
        Mapel=mapel,
        Semester=semester,
        Tahun_Pelajaran=tahun_pelajaran
    ).first()

    if not seting_soal:
        messages.error(request, "Seting Soal tidak ditemukan.")
        return redirect("cbt:daftar_nilai")

    # Hitung total nilai maksimal
    soal_list = models.Soal_Siswa.objects.filter(Kode_Soal=seting_soal)
    total_nilai_maksimal = sum(soal.Nilai or 0 for soal in soal_list)
    

    jawaban = models.Answer.objects.filter(
        Kode_Soal=seting_soal,
        Kelas=kelas,
        Rombel=rombel
    ).select_related("Nama_User", "Nomor_Soal")

    nilai_siswa = {}
    for ans in jawaban:
        user = ans.Nama_User
        if user.id not in nilai_siswa:
            nama = user.get_full_name() or getattr(user, 'Nama', None) or user.username
            nilai_siswa[user.id] = {
                "nama": nama,
                "total_nilai": 0,
                "total_benar": 0,
                "total_soal": 0
            }

        # Jika ada jawaban, tambah total soal
        if ans.Jawaban is not None:
            nilai_siswa[user.id]["total_soal"] += 1

        # Jika benar, tambah nilai sesuai bobot soal
        if ans.Jawaban_Benar:
            nilai_soal = ans.Nomor_Soal.Nilai or 0
            nilai_siswa[user.id]["total_nilai"] += nilai_soal
            nilai_siswa[user.id]["total_benar"] += 1

    # Hitung persentase dan nilai akhir
    for data in nilai_siswa.values():
        total_soal = data['total_soal']
        data["persentase"] = round((data["total_benar"] / total_soal) * 100, 2) if total_soal > 0 else 0
        data["nilai_akhir"] = round((data["total_nilai"] / total_nilai_maksimal) * 100, 2) if total_nilai_maksimal > 0 else 0

    html_string = render_to_string("staff/nilai_pdf.html", {
        "mapel": mapel,
        "kelas": kelas,
        "rombel": rombel,
        "nilai_siswa": sorted(nilai_siswa.values(), key=lambda x: x["nama"]),
    })



    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'filename=Daftar_Nilai_{mapel}_{kelas}_{rombel}.pdf'

    with tempfile.NamedTemporaryFile(delete=True) as output:
        HTML(string=html_string).write_pdf(output.name)
        output.seek(0)
        response.write(output.read())

    return response







@login_required(login_url=settings.LOGIN_URL)
@user_passes_test(lambda user: user.is_staff, login_url=settings.LOGIN_URL)
@csrf_protect
def arsip_soal (request):
    cari = request.POST.get('cari', request.GET.get('cari', '')) 
    items_per_page = request.GET.get('items_per_page', '30')  # Ubah ke string dulu
    
    # Jika memilih "Semua", set items_per_page ke jumlah besar
    if items_per_page == 'all':
        items_per_page = 1000000  # Angka besar untuk menampilkan semua
    else:
        try:
            items_per_page = int(items_per_page)
        except (ValueError, TypeError):
            items_per_page = 30  # Default jika invalid

    data_list = models.TahunPelajaran.objects.filter(status=False)
    if cari:
        data_list = data_list.filter(Tahun_Pelajaran=cari)
    paginator = Paginator(data_list, items_per_page)
    page_number = request.GET.get('page', 1)
    
    try:
        Data = paginator.page(page_number)
    except PageNotAnInteger:
        Data = paginator.page(1)
    except EmptyPage:
        Data = paginator.page(paginator.num_pages)

    semua = str(items_per_page) if items_per_page != 1000000 else 'all'
    Data.start_index = (Data.start_index() - 1)
    jumlah = data_list.count()
    
    context = {
        "data":"Arsip Soal",
        "judul":"CBT-arsip Soal",
        "Data":Data,
        "jumlah": jumlah,
        "cari": cari,
        "items_per_page": semua ,
        "lembaga": "Arsip Soal",
        "placeholder": "Tahun Pelajaran",
        "icon":"bi bi-database-check"
    }
    
    
    return render(request, "staff/arsip_soal.html", context)




@login_required(login_url=settings.LOGIN_URL)
@user_passes_test(lambda user: user.is_staff, login_url=settings.LOGIN_URL)
@csrf_protect
def setting_soal_arsip(request, pk):
    cari = request.POST.get('cari', request.GET.get('cari', ''))
    items_per_page = request.GET.get('items_per_page', '30')

    if items_per_page == 'all':
        items_per_page = 1000000
    else:
        try:
            items_per_page = int(items_per_page)
        except (ValueError, TypeError):
            items_per_page = 30

    # Ambil tahun pelajaran berdasarkan pk
    tahun_aktif = get_object_or_404(models.TahunPelajaran, id=pk)

    # Ambil semester dari relasi tahun pelajaran
    semester_obj = tahun_aktif.semester
    semester = semester_obj.Semester if semester_obj else None

    if semester and tahun_aktif:
        data_list = models.SetingSoal.objects.filter(
            Semester=semester,  # Ini string, sesuai model SetingSoal
            Tahun_Pelajaran=tahun_aktif
        )
    else:
        data_list = models.SetingSoal.objects.none()

    if cari:
        data_list = data_list.filter(Mapel__Nama_Mapel__icontains=cari)

    # Pagination
    paginator = Paginator(data_list, items_per_page)
    page_number = request.GET.get('page', 1)

    try:
        Data = paginator.page(page_number)
    except PageNotAnInteger:
        Data = paginator.page(1)
    except EmptyPage:
        Data = paginator.page(paginator.num_pages)

    Data.start_index = (Data.start_index() - 1)
    jumlah = data_list.count()

    context = {
        "data": f"Soal Semester {semester} TP.{tahun_aktif.Tahun_Pelajaran}",
        "judul": "CBT-Setting",
        "Data": Data,
        "icon": "bi bi-database-check",
        "semester": semester,
        "jumlah": jumlah,
        "cari": cari,
        "items_per_page": str(items_per_page) if items_per_page != 1000000 else 'all',
        "lembaga": "Mapel Kelas",
        "placeholder": "Mapel",
    }

    return render(request, 'staff/seting_soal_arsip.html', context)








@login_required(login_url=settings.LOGIN_URL)
@user_passes_test(lambda user: user.is_staff, login_url=settings.LOGIN_URL)
@csrf_protect
def salin_setingsoal(request, pk):
    # Ambil SetingSoal lama
    soal_lama = get_object_or_404(models.SetingSoal, pk=pk)

    # Ambil semester dan tahun pelajaran aktif
    semester_aktif = models.SEMESTER.objects.filter( status=True).first()
    tahun_aktif = models.TahunPelajaran.objects.filter(status=True).first()
    tahun_tidak_Active = models.TahunPelajaran.objects.filter(status=False).first()

    if not semester_aktif or not tahun_aktif:
        messages.error(
            request,
            f"Semester atau Tahun Pelajaran aktif tidak ditemukan. "
            f"(Semester: {semester_aktif}, Tahun: {tahun_aktif})"
        )
        return redirect(reverse('cbt:setting_soal_arsip', kwargs={'pk': tahun_tidak_Active.pk}))

    try:
        with transaction.atomic():
            # Salin SetingSoal baru dengan kode soal baru
            soal_baru = models.SetingSoal.objects.create(
                Nama_User=request.user,
                Nama_Lembaga=soal_lama.Nama_Lembaga,
                Kelas=soal_lama.Kelas,
                Mapel=soal_lama.Mapel,
                Semester=semester_aktif.Semester,
                Tahun_Pelajaran=tahun_aktif,
                durasi_menit=soal_lama.durasi_menit,
                aktif=False,
                arsip_soal=False,
                # Kode_Soal otomatis dibuat dari default=buat_nomor_baru
            )

            # Salin semua Soal_Siswa dan hubungkan dengan soal_baru
            soal_siswa_lama = models.Soal_Siswa.objects.filter(Kode_Soal=soal_lama)
            for s in soal_siswa_lama:
                models.Soal_Siswa.objects.create(
                    Nama_User=request.user,
                    Kode_Soal=soal_baru,
                    Mapel = s.Mapel,
                    Kelas = s.Kelas,
                    Nomor = s.Nomor,
                    Soal=s.Soal,
                    A=s.A,
                    B=s.B,
                    C=s.C,
                    D=s.D,
                    Kunci_Jawaban=s.Kunci_Jawaban,
                    Nilai = s.Nilai
                    # Tambahkan field lain jika ada
                )

            messages.success(
                request,
                f"Berhasil menyalin soal {soal_lama.Mapel} kode sola :{soal_lama.Kode_Soal} Menjadi '{soal_baru.Kode_Soal}"
            )

    except Exception as e:
        messages.error(request, f"Gagal menyalin soal: {str(e)}")

    return redirect(reverse('cbt:setting_soal_arsip', kwargs={'pk': tahun_tidak_Active.pk}))
    


